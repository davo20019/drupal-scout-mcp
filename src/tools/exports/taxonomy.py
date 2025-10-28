"""
Taxonomy export tool for Drupal Scout MCP.

Provides CSV and Excel export functionality for taxonomy terms:
- export_taxonomy_usage_to_csv: Export taxonomy terms to CSV with usage analysis
- export_taxonomy_usage_to_excel: Export taxonomy terms to Excel with full node details and multi-tab support

This tool bypasses MCP token limits by writing directly to filesystem.
"""

import csv
import json
import logging
import subprocess
from datetime import datetime
from pathlib import Path
from typing import List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import from core modules
from src.core.drush import run_drush_command

# Import shared export utilities
from src.tools.exports.common import get_export_config, validate_export_path

# Import MCP instance from server
from server import mcp

# Get logger
logger = logging.getLogger(__name__)


@mcp.tool()
def export_taxonomy_usage_to_csv(
    vocabulary: str,
    output_path: Optional[str] = None,
    check_code: bool = False,
    summary_only: bool = True,
    limit: int = 0,
) -> str:
    """
    Export ALL taxonomy term usage data directly to a CSV file, bypassing MCP token limits.

    **BEST SOLUTION FOR LARGE VOCABULARIES (500+ terms)**

    This tool writes CSV directly to the filesystem, avoiding MCP's 25K token response limit.
    Perfect for vocabularies with hundreds or thousands of terms.

    **How It Works:**
    1. Fetches ALL terms from vocabulary (no limit)
    2. Writes CSV directly to file system
    3. Returns file path and summary stats (tiny response)
    4. LLM can then tell user where file is saved

    **When to Use This Tool:**
    - User asks to "export to CSV" or "save to file"
    - Vocabulary has 200+ terms
    - User wants complete dataset without truncation
    - Avoiding token limit issues

    **Workflow Example:**
    ```
    User: "Export all tags to CSV"
    LLM: export_taxonomy_usage_to_csv(vocabulary="tags", output_path="/tmp/tags_export.csv")
    Response: "✅ Exported 751 terms to /tmp/tags_export.csv (45 KB)"
    LLM: "I've exported all 751 terms to /tmp/tags_export.csv. The file includes..."
    ```

    Args:
        vocabulary: Vocabulary machine name (e.g., "topics", "tags")
        output_path: Optional CSV file path. If not provided, auto-generates in Drupal root:
                     {drupal_root}/taxonomy_export_{vocabulary}_{timestamp}.csv
        check_code: If True, scans custom code for term references (slower).
                    Default: False (recommended for large exports)
                    Only applies when summary_only=False
        summary_only: If True, exports minimal columns (tid, name, count, needs_check).
                      If False, exports FULL details including:
                        - tid, name, description, parent, children
                        - content_count (total usage)
                        - content_usage_sample (first 5 nodes using this term)
                        - fields_with_usage (which fields use this term)
                        - code_usage (hardcoded references in custom code, if check_code=True)
                        - config_usage (config file references, if check_code=True)
                        - safe_to_delete (YES/NO based on all checks)
                        - warnings (reasons why term might not be safe to delete)
                      Default: True (faster for large vocabularies)
        limit: Max terms to export. 0 = all terms. Default: 0

    Returns:
        JSON with file path, stats, and preview:
        {
            "file_path": "/tmp/tags_export.csv",
            "total_terms": 751,
            "file_size_kb": 45,
            "columns": ["tid", "name", "count", "safe_to_delete"],
            "preview": "First 3 rows preview..."
        }

    **Advantages Over get_all_taxonomy_usage():**
    - No token limits (can handle 10,000+ terms)
    - Faster (no JSON serialization overhead)
    - File persists for user to download/analyze
    - Response is tiny regardless of term count

    **Performance:**
    - 100 terms: ~5 seconds
    - 500 terms: ~20 seconds
    - 1000 terms: ~40 seconds
    - 5000 terms: ~3 minutes
    """
    try:
        # Get config and verify database connection
        success, result, msg = get_export_config()
        if not success:
            return result["error_response"]

        drupal_root = result["drupal_root"]

        # Auto-generate path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Save in Drupal root directory for easy access in IDE
            output_path = str(drupal_root / f"taxonomy_export_{vocabulary}_{timestamp}.csv")

        # Validate path is within safe boundaries
        is_valid, error_msg = validate_export_path(output_path, drupal_root)
        if not is_valid:
            return json.dumps({"_error": True, "message": f"Invalid export path: {error_msg}"})

        # Validate path is writable
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Get ALL terms directly from helper function (no limit)
        logger.info(f"Fetching all terms from vocabulary '{vocabulary}'...")

        # In full mode, fetch sample nodes; in summary mode, skip them for speed
        max_samples = 0 if summary_only else 5

        terms = _get_all_terms_usage_from_drush(
            vocabulary=vocabulary,
            max_sample_nodes=max_samples,
            summary_only=summary_only,
            limit=limit,
        )

        if not terms:
            return json.dumps(
                {
                    "_error": True,
                    "message": f"Could not fetch terms for vocabulary '{vocabulary}'",
                    "suggestion": "Verify vocabulary machine name is correct.",
                }
            )

        total_terms = len(terms)

        # Add code/config usage if requested
        if check_code and not summary_only:
            terms = _add_code_usage_to_terms(terms, drupal_root)

        # Determine CSV columns based on summary_only mode
        if summary_only:
            columns = ["tid", "name", "count", "needs_check"]
        else:
            # Full mode - include ALL available details
            columns = [
                "tid",
                "name",
                "description",
                "parent",
                "children",
                "content_count",
                "content_usage_sample",  # Sample nodes (first 5)
                "fields_with_usage",
                "code_usage",
                "config_usage",
                "safe_to_delete",
                "warnings",
            ]

        # Write CSV
        logger.info(f"Writing {total_terms} terms to {output_path}...")
        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()

            for term in terms:
                row = {}
                for col in columns:
                    value = term.get(col, "")

                    # Special handling for different field types
                    if col == "content_usage_sample":
                        # Get sample nodes from content_usage array
                        content_usage = term.get("content_usage", [])
                        if content_usage:
                            samples = [
                                f"nid:{item.get('nid')} ({item.get('title', 'Untitled')})"
                                for item in content_usage[:5]
                            ]
                            value = " | ".join(samples)
                        else:
                            value = ""

                    elif col == "content_count":
                        # Use content_count if available, otherwise count content_usage
                        value = term.get("content_count") or term.get("count", 0)

                    elif col == "code_usage":
                        # Format code references
                        code_usage = term.get("code_usage", [])
                        if code_usage:
                            refs = [
                                f"{item.get('file', '')}:{item.get('line', '')}"
                                for item in code_usage
                            ]
                            value = " | ".join(refs)
                        else:
                            value = ""

                    elif col == "config_usage":
                        # Format config references
                        config_usage = term.get("config_usage", [])
                        if config_usage:
                            refs = [
                                f"{item.get('config_name', '')} ({item.get('config_type', '')})"
                                for item in config_usage
                            ]
                            value = " | ".join(refs)
                        else:
                            value = ""

                    elif isinstance(value, list):
                        # Handle any other lists/arrays
                        value = " | ".join(str(v) for v in value)

                    elif isinstance(value, bool):
                        # Convert boolean to readable text
                        value = "YES" if value else "NO"

                    row[col] = value

                writer.writerow(row)

        # Get file size
        file_size_kb = round(output_file.stat().st_size / 1024, 1)

        # Create preview (first 3 rows)
        with open(output_path, "r", encoding="utf-8") as f:
            preview_lines = [f.readline().strip() for _ in range(4)]  # Header + 3 rows
            preview = "\n".join(preview_lines)

        result = {
            "success": True,
            "file_path": str(output_path),
            "total_terms": total_terms,
            "file_size_kb": file_size_kb,
            "columns": columns,
            "preview": preview,
            "message": f"✅ Successfully exported {total_terms} terms to {output_path} ({file_size_kb} KB)",
        }

        return json.dumps(result, indent=2)

    except Exception as e:
        logger.error(f"Error in export_taxonomy_usage_to_csv: {e}", exc_info=True)
        return json.dumps({"_error": True, "message": str(e)})


@mcp.tool()
def export_taxonomy_usage_to_excel(
    vocabulary: str,
    output_path: Optional[str] = None,
    include_formatting: bool = True,
    full_details: bool = False,
    check_code: bool = False,
    limit: int = 0,
) -> str:
    """
    Export taxonomy term usage data to Excel with merged cell layout showing all pages per term.

    **BEST FOR DETAILED TAXONOMY ANALYSIS WITH NODE INFORMATION**

    This tool creates an Excel workbook with a single sheet layout where:
    - Each term occupies a merged cell spanning all its pages (rows)
    - Each page gets its own row with complete node details
    - Easy to scan which pages belong to which term

    **Layout Example:**
    ```
    | Term Name    | Node ID | Title         | Status    | URL         | Author | ... |
    |--------------|---------|---------------|-----------|-------------|--------|-----|
    | Technology   | 123     | AI Article    | published | /ai-article | John   | ... |
    |      ↓       | 124     | Web Dev       | published | /web-dev    | Jane   | ... |
    |      ↓       | 125     | Cloud Guide   | published | /cloud      | Bob    | ... |
    | Business     | 201     | Marketing     | published | /marketing  | Alice  | ... |
    |      ↓       | 202     | Sales Tips    | published | /sales      | Tom    | ... |
    ```

    **Key Features:**
    - Single sheet layout (no tabs) - handles thousands of pages easily
    - Term name in merged cell spanning all its pages for easy visual grouping
    - Full node details for EVERY page: title, status, type, url_alias, canonical_url,
      author, dates, taxonomy terms, entity references, metatags, revision count
    - Professional formatting: Bold headers, auto-sized columns, freeze panes, filters
    - Terms without pages show "(No pages)" in the first row

    **When to Use:**
    - User asks to "export to Excel" or wants detailed page information
    - Need to see exactly which pages use each taxonomy term
    - Want full node metadata (status, path, author, etc.) for each page
    - Need easy visual grouping by term

    Args:
        vocabulary: Vocabulary machine name (e.g., "topics", "tags")
        output_path: Optional Excel file path (.xlsx). If not provided, auto-generates in Drupal root:
                     {drupal_root}/taxonomy_export_{vocabulary}_{timestamp}.xlsx
        include_formatting: If True, applies professional formatting (headers, colors, column widths).
                          Default: True
        full_details: If False (default), exports essential node columns (ID, Title, Type, Status, URL Alias).
                     If True, exports ALL node details (20+ columns including author, dates, taxonomy terms,
                     entity references, metatags, revisions, language, promote/sticky flags).
                     Default: False (quick/essential mode)
        check_code: If True, scans custom code for term references (slower).
                   Default: False
        limit: Max terms to export. 0 = all terms. Default: 0

    Returns:
        JSON with file path, stats, and summary:
        {
            "success": true,
            "file_path": "/path/to/taxonomy_export_tags_20251027.xlsx",
            "total_terms": 156,
            "total_pages": 2341,
            "total_rows": 2341,
            "file_size_kb": 450.5,
            "message": "✅ Successfully exported..."
        }

    **Node Details Included for Each Page:**
    - nid: Node ID
    - title: Page title
    - type: Content type
    - status: published/unpublished
    - url_alias: Internal path (/about-us)
    - canonical_url: Full URL
    - author: Author name
    - author_uid: Author user ID
    - created: Creation date
    - changed: Last modified date
    - langcode: Language
    - promote: Promoted to front page
    - sticky: Sticky at top of lists
    - taxonomy_terms: All taxonomy terms on this node (all vocabularies)
    - entity_references: Other entity references (media, nodes, etc.)
    - metatag_title: SEO title (if metatag module enabled)
    - metatag_description: SEO description
    - revision_count: Number of revisions

    **Excel Structure (Single Sheet):**
    - Column A: Term Name (merged across all page rows)
    - Column B: Term Description (merged across all page rows)
    - Columns C+: Node details (one row per page)

    **Performance:**
    - 100 terms, 500 pages: ~15 seconds
    - 500 terms, 2000 pages: ~60 seconds
    - 1000 terms, 5000 pages: ~2 minutes
    - Can handle tens of thousands of rows (Excel limit: 1,048,576 rows)

    **Requirements:**
    - Requires openpyxl library: pip install openpyxl
    - Will return error if openpyxl not installed

    **Example Usage:**
    ```
    # Basic export (quick mode - essential columns only)
    export_taxonomy_usage_to_excel(vocabulary="tags")
    # Exports: Term Name, Term Description, Node ID, Node Title, Node Type, Node Status, Node URL Alias

    # Full details mode (all 20+ columns)
    export_taxonomy_usage_to_excel(vocabulary="tags", full_details=True)
    # Adds: Canonical URL, Author, Dates, Language, Taxonomy Terms, Entity Refs, Metatags, etc.

    # Include code scanning
    export_taxonomy_usage_to_excel(vocabulary="categories", check_code=True)

    # Limit to first 100 terms for testing
    export_taxonomy_usage_to_excel(vocabulary="tags", limit=100)
    ```
    """
    try:
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            return json.dumps({
                "_error": True,
                "message": "openpyxl library not installed. Install with: pip install openpyxl",
                "suggestion": "Run: pip install openpyxl"
            })

        # Get config and verify database connection
        success, result, msg = get_export_config()
        if not success:
            return result["error_response"]

        drupal_root = result["drupal_root"]

        # Auto-generate path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(drupal_root / f"taxonomy_export_{vocabulary}_{timestamp}.xlsx")

        # Validate path is within safe boundaries
        is_valid, error_msg = validate_export_path(output_path, drupal_root)
        if not is_valid:
            return json.dumps({"_error": True, "message": f"Invalid export path: {error_msg}"})

        # Validate path is writable and has .xlsx extension
        output_file = Path(output_path)
        if output_file.suffix.lower() != '.xlsx':
            output_path = str(output_file.with_suffix('.xlsx'))
            output_file = Path(output_path)

        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Get ALL terms with FULL node details
        logger.info(f"Fetching all terms with full node details from vocabulary '{vocabulary}'...")

        terms = _get_all_terms_with_full_nodes_from_drush(
            vocabulary=vocabulary,
            limit=limit,
        )

        if not terms:
            return json.dumps({
                "_error": True,
                "message": f"Could not fetch terms for vocabulary '{vocabulary}'",
                "suggestion": "Verify vocabulary machine name is correct.",
            })

        total_terms = len(terms)

        # Add code/config usage if requested
        if check_code:
            terms = _add_code_usage_to_terms(terms, drupal_root)

        # Sort terms by page count (descending) - terms with pages first, then terms without pages
        terms = sorted(terms, key=lambda t: len(t.get('nodes', [])), reverse=True)

        # Create Excel workbook with single sheet
        logger.info(f"Creating Excel workbook with {total_terms} terms...")
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Taxonomy Export"

        # Define columns based on detail level
        if full_details:
            # Full mode: All node details
            columns = [
                "Term ID",
                "Term Name",
                "Term Description",
                "Node ID",
                "Node Title",
                "Node Type",
                "Node Status",
                "Node URL Alias",
                "Node Canonical URL",
                "Node Author",
                "Node Author UID",
                "Node Created",
                "Node Changed",
                "Node Language",
                "Node Promote",
                "Node Sticky",
                "Node Taxonomy Terms",
                "Node Entity References",
                "Node Metatag Title",
                "Node Metatag Description",
                "Node Revision Count"
            ]
        else:
            # Quick mode: Essential columns only (default)
            columns = [
                "Term ID",
                "Term Name",
                "Term Description",
                "Node ID",
                "Node Title",
                "Node Type",
                "Node Status",
                "Node URL Alias"
            ]

        # Write header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=col_name)
            if include_formatting:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Freeze header row
        if include_formatting:
            sheet.freeze_panes = "A2"

        # Track statistics
        current_row = 2
        total_pages = 0

        # Process each term
        for term in terms:
            tid = term.get("tid")
            term_name = term.get("name", "")
            term_description = term.get("description", "")
            nodes = term.get("nodes", [])
            node_count = len(nodes)
            total_pages += node_count

            # If term has no pages, still show it in one row
            if node_count == 0:
                # Write term with no pages
                sheet.cell(row=current_row, column=1, value=tid)
                sheet.cell(row=current_row, column=2, value=term_name)
                sheet.cell(row=current_row, column=3, value=term_description)
                sheet.cell(row=current_row, column=4, value="(No pages)")

                if include_formatting:
                    sheet.cell(row=current_row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
                    sheet.cell(row=current_row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
                    sheet.cell(row=current_row, column=3).alignment = Alignment(vertical="top", wrap_text=True)
                    sheet.cell(row=current_row, column=4).alignment = Alignment(vertical="center")
                    sheet.cell(row=current_row, column=4).font = Font(italic=True, color="808080")
                    # Light gray background for no-pages row
                    for col in range(1, len(columns) + 1):
                        sheet.cell(row=current_row, column=col).fill = PatternFill(
                            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
                        )

                current_row += 1
            else:
                # Term has pages - write one row per page with merged term cells
                start_row = current_row

                for node in nodes:
                    # Term ID, name and description in first three columns
                    sheet.cell(row=current_row, column=1, value=tid)
                    sheet.cell(row=current_row, column=2, value=term_name)
                    sheet.cell(row=current_row, column=3, value=term_description)

                    # Node details in remaining columns
                    if full_details:
                        # Full mode: All details
                        node_data = [
                            node.get('nid'),
                            node.get('title'),
                            node.get('type'),
                            node.get('status'),
                            node.get('url_alias'),
                            node.get('canonical_url'),
                            node.get('author'),
                            node.get('author_uid'),
                            datetime.fromtimestamp(node.get('created', 0)).strftime("%Y-%m-%d %H:%M:%S") if node.get('created') else '',
                            datetime.fromtimestamp(node.get('changed', 0)).strftime("%Y-%m-%d %H:%M:%S") if node.get('changed') else '',
                            node.get('langcode'),
                            "YES" if node.get('promote') else "NO",
                            "YES" if node.get('sticky') else "NO",
                            " | ".join(node.get('taxonomy_terms', [])),
                            " | ".join(node.get('entity_references', [])),
                            node.get('metatag_title'),
                            node.get('metatag_description'),
                            node.get('revision_count')
                        ]
                    else:
                        # Quick mode: Essential fields only
                        node_data = [
                            node.get('nid'),
                            node.get('title'),
                            node.get('type'),
                            node.get('status'),
                            node.get('url_alias')
                        ]

                    for col_idx, value in enumerate(node_data, start=4):
                        cell = sheet.cell(row=current_row, column=col_idx, value=value)
                        if include_formatting:
                            cell.alignment = Alignment(vertical="top", wrap_text=True)

                    current_row += 1

                # Merge term ID, name and description cells across all rows for this term
                end_row = current_row - 1
                if include_formatting and end_row > start_row:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                    sheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

                    # Center align and apply formatting to merged cells
                    sheet.cell(row=start_row, column=1).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    sheet.cell(row=start_row, column=2).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                    sheet.cell(row=start_row, column=3).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

                    # Bold the term name
                    sheet.cell(row=start_row, column=2).font = Font(bold=True)

        # Auto-size columns
        if include_formatting:
            _auto_size_columns(sheet, columns)

        # Add autofilter
        if include_formatting and total_pages > 0:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{current_row - 1}"

        # Save workbook
        logger.info(f"Saving Excel workbook to {output_path}...")
        wb.save(output_path)

        # Get file size
        file_size_kb = round(output_file.stat().st_size / 1024, 1)

        result = {
            "success": True,
            "file_path": str(output_path),
            "total_terms": total_terms,
            "total_pages": total_pages,
            "total_rows": current_row - 2,  # Subtract header and start from 0
            "file_size_kb": file_size_kb,
            "message": f"✅ Successfully exported {total_terms} terms with {total_pages} total pages to {output_path} ({file_size_kb} KB)\n"
                      f"   Single sheet with {current_row - 2} rows (merged cells for terms with multiple pages)"
        }

        return json.dumps(result, indent=2)

    except Exception as e:
        logger.error(f"Error in export_taxonomy_usage_to_excel: {e}", exc_info=True)
        return json.dumps({"_error": True, "message": str(e)})


def _get_all_terms_usage_from_drush(
    vocabulary: str, max_sample_nodes: int = 5, summary_only: bool = False, limit: int = 0
) -> Optional[List[dict]]:
    """Get usage data for all terms in a vocabulary via optimized drush query."""
    try:
        php_code = f"""
        $vocabulary = '{vocabulary}';
        $max_samples = {max_sample_nodes};
        $summary_only = {str(summary_only).lower()};
        $limit = {limit};

        $term_storage = \\Drupal::entityTypeManager()->getStorage('taxonomy_term');
        $vocab_storage = \\Drupal::entityTypeManager()->getStorage('taxonomy_vocabulary');
        $database = \\Drupal::database();
        $node_storage = \\Drupal::entityTypeManager()->getStorage('node');

        // Load vocabulary
        $vocab = $vocab_storage->load($vocabulary);
        if (!$vocab) {{
            echo json_encode(null);
            exit;
        }}

        // Load all terms in vocabulary
        $query = \\Drupal::entityQuery('taxonomy_term')
            ->accessCheck(FALSE)
            ->condition('vid', $vocabulary);

        // Apply limit if specified
        if ($limit > 0) {{
            $query->range(0, $limit);
        }}

        $term_ids = $query->execute();

        if (empty($term_ids)) {{
            echo json_encode([]);
            exit;
        }}

        $terms = $term_storage->loadMultiple($term_ids);
        $results = [];
        $total_terms = count($terms);

        // Progress: Initial status
        fwrite(STDERR, "Analyzing " . $total_terms . " terms in vocabulary '" . $vocabulary . "'...\\n");

        // Get field map once (reuse for all terms)
        $field_map = \\Drupal::service('entity_field.manager')->getFieldMapByFieldType('entity_reference');

        // Build list of field tables to check
        $field_tables = [];
        foreach ($field_map['node'] ?? [] as $field_name => $field_info) {{
            $table_name = 'node__' . $field_name;
            $column_name = $field_name . '_target_id';

            // Test if table exists
            try {{
                $database->query("SELECT 1 FROM {{" . $table_name . "}} LIMIT 1")->fetchField();
                $field_tables[] = [
                    'table' => $table_name,
                    'column' => $column_name,
                    'field_name' => $field_name
                ];
            }} catch (\\Exception $e) {{
                // Table doesn't exist - skip
            }}
        }}

        fwrite(STDERR, "Checking " . count($field_tables) . " field tables for term usage...\\n");

        $processed = 0;
        foreach ($terms as $term) {{
            $processed++;

            // Progress: Show every 50 terms or at key milestones
            if ($processed % 50 === 0 || $processed === 1 || $processed === $total_terms) {{
                fwrite(STDERR, "Progress: " . $processed . "/" . $total_terms . " terms processed\\n");
            }}
            $tid = (int)$term->id();

            // Get term metadata
            $parents = $term_storage->loadParents($tid);
            $parent_name = $parents ? reset($parents)->getName() : null;

            $children = $term_storage->loadChildren($tid);
            $children_names = array_map(function($child) {{ return $child->getName(); }}, $children);

            // Find content using this term (optimized)
            $nids_found = [];
            $fields_found_usage = [];

            // Check taxonomy_index first (fast)
            $taxonomy_index_nids = $database->query(
                "SELECT DISTINCT nid FROM {{taxonomy_index}} WHERE tid = :tid",
                [':tid' => $tid]
            )->fetchCol();

            if ($taxonomy_index_nids) {{
                $nids_found = array_merge($nids_found, $taxonomy_index_nids);
            }}

            // Check field data tables
            foreach ($field_tables as $field_info) {{
                try {{
                    $field_nids = $database->query(
                        "SELECT DISTINCT entity_id FROM {{" . $field_info['table'] . "}} WHERE " . $field_info['column'] . " = :tid",
                        [':tid' => $tid]
                    )->fetchCol();

                    if ($field_nids) {{
                        $nids_found = array_merge($nids_found, $field_nids);
                        $fields_found_usage[] = $field_info['field_name'];
                    }}
                }} catch (\\Exception $e) {{
                    // Query failed - skip
                }}
            }}

            // Get unique NIDs
            $nids_found = array_unique($nids_found);
            $total_content_count = count($nids_found);

            // Determine if safe to delete
            $safe_to_delete = ($total_content_count === 0 && empty($children_names));

            // Build result based on summary_only flag
            if ($summary_only) {{
                // Ultra-compact mode: just essentials
                // Note: 'needs_check' indicates this is content-only analysis
                $results[] = [
                    'tid' => $tid,
                    'name' => $term->getName(),
                    'count' => $total_content_count,
                    'needs_check' => $safe_to_delete  // TRUE = candidate for deletion, needs full check
                ];
            }} else {{
                // Full mode: get sample nodes
                $sample_nids = array_slice($nids_found, 0, $max_samples);
                $content_usage = [];

                if (!empty($sample_nids)) {{
                    $nodes = $node_storage->loadMultiple($sample_nids);
                    foreach ($nodes as $node) {{
                        $content_usage[] = [
                            'nid' => (int)$node->id(),
                            'title' => $node->getTitle(),
                            'type' => $node->bundle()
                        ];
                    }}
                }}

                $warnings = [];
                if ($total_content_count > 0) {{
                    $warnings[] = "Used in " . $total_content_count . " content items";
                }}
                if (!empty($children_names)) {{
                    $warnings[] = "Has " . count($children_names) . " child terms";
                }}

                $results[] = [
                    'tid' => $tid,
                    'name' => $term->getName(),
                    'description' => $term->getDescription(),
                    'vocabulary' => $vocabulary,
                    'vocabulary_label' => $vocab->label(),
                    'parent' => $parent_name,
                    'children' => array_values($children_names),
                    'content_count' => $total_content_count,
                    'content_usage' => $content_usage,
                    'fields_with_usage' => $fields_found_usage,
                    'safe_to_delete' => $safe_to_delete,
                    'warnings' => $warnings
                ];
            }}
        }}

        // Progress: Completion
        fwrite(STDERR, "Analysis complete! Processed " . $total_terms . " terms.\\n");

        echo json_encode($results);
        """

        result = run_drush_command(["ev", php_code.strip()], timeout=120)
        return result if result and isinstance(result, list) else None

    except Exception as e:
        logger.debug(f"Could not get all terms usage from drush: {e}")
        return None


def _add_code_usage_to_terms(terms_data: List[dict], drupal_root: Path) -> List[dict]:
    """Scan custom code and config for hardcoded term references."""
    # Build lookup maps for efficient searching
    tid_to_term = {term["tid"]: term for term in terms_data}
    name_to_terms = {}
    for term in terms_data:
        name = term["name"]
        if name not in name_to_terms:
            name_to_terms[name] = []
        name_to_terms[name].append(term)

    # Initialize code_usage and config_usage for all terms
    for term in terms_data:
        term["code_usage"] = []
        term["config_usage"] = []

    # Scan custom modules directory
    custom_dir = drupal_root / "web" / "modules" / "custom"
    if not custom_dir.exists():
        custom_dir = drupal_root / "modules" / "custom"

    if custom_dir.exists():
        # Scan for term IDs in PHP files
        try:
            # Search for ->load(TID) or Term::load(TID)
            for tid in tid_to_term.keys():
                result = subprocess.run(
                    ["grep", "-r", "-n", f"load({tid})", str(custom_dir)],
                    capture_output=True,
                    text=True,
                    timeout=10,
                )
                if result.returncode == 0:
                    for line in result.stdout.strip().split("\n"):
                        if line:
                            parts = line.split(":", 2)
                            if len(parts) >= 3:
                                file_path = parts[0].replace(str(drupal_root) + "/", "")
                                line_num = parts[1]
                                context = parts[2].strip()[:100]
                                tid_to_term[tid]["code_usage"].append(
                                    {"file": file_path, "line": line_num, "context": context}
                                )
                                if "Referenced in custom code" not in tid_to_term[tid]["warnings"]:
                                    tid_to_term[tid]["warnings"].append("Referenced in custom code")
                                    tid_to_term[tid]["safe_to_delete"] = False
        except Exception as e:
            logger.debug(f"Error scanning for term IDs: {e}")

        # Scan for term names in PHP/Twig files
        try:
            for name, terms_list in name_to_terms.items():
                # Escape single quotes for grep
                escaped_name = name.replace("'", "\\'")
                result = subprocess.run(
                    ["grep", "-r", "-n", f"'{escaped_name}'", str(custom_dir)],
                    capture_output=True,
                    text=True,
                    timeout=10,
                )
                if result.returncode == 0:
                    for line in result.stdout.strip().split("\n")[
                        :5
                    ]:  # Limit to 5 matches per term
                        if line:
                            parts = line.split(":", 2)
                            if len(parts) >= 3:
                                file_path = parts[0].replace(str(drupal_root) + "/", "")
                                line_num = parts[1]
                                context = parts[2].strip()[:100]
                                for term in terms_list:
                                    term["code_usage"].append(
                                        {"file": file_path, "line": line_num, "context": context}
                                    )
                                    if "Referenced in custom code" not in term["warnings"]:
                                        term["warnings"].append("Referenced in custom code")
                                        term["safe_to_delete"] = False
        except Exception as e:
            logger.debug(f"Error scanning for term names: {e}")

    # Scan config directory for term references
    config_dir = drupal_root / "config" / "sync"
    if not config_dir.exists():
        config_dir = drupal_root / "web" / "sites" / "default" / "files" / "config"

    if config_dir.exists():
        try:
            for tid in tid_to_term.keys():
                result = subprocess.run(
                    ["grep", "-r", "-l", f": {tid}", str(config_dir)],
                    capture_output=True,
                    text=True,
                    timeout=10,
                )
                if result.returncode == 0:
                    for config_file in result.stdout.strip().split("\n"):
                        if config_file:
                            config_name = Path(config_file).stem
                            tid_to_term[tid]["config_usage"].append(
                                {"config_name": config_name, "config_type": "yaml"}
                            )
                            if "Referenced in config files" not in tid_to_term[tid]["warnings"]:
                                tid_to_term[tid]["warnings"].append("Referenced in config files")
                                tid_to_term[tid]["safe_to_delete"] = False
        except Exception as e:
            logger.debug(f"Error scanning config files: {e}")

    return terms_data


def _get_all_terms_with_full_nodes_from_drush(
    vocabulary: str, limit: int = 0
) -> Optional[List[dict]]:
    """Get all terms with complete node details for each term's usage."""
    try:
        php_code = f"""
        $vocabulary = '{vocabulary}';
        $limit = {limit};

        $term_storage = \\Drupal::entityTypeManager()->getStorage('taxonomy_term');
        $vocab_storage = \\Drupal::entityTypeManager()->getStorage('taxonomy_vocabulary');
        $database = \\Drupal::database();
        $node_storage = \\Drupal::entityTypeManager()->getStorage('node');

        // Load vocabulary
        $vocab = $vocab_storage->load($vocabulary);
        if (!$vocab) {{
            echo json_encode(null);
            exit;
        }}

        // Load all terms in vocabulary
        $query = \\Drupal::entityQuery('taxonomy_term')
            ->accessCheck(FALSE)
            ->condition('vid', $vocabulary);

        if ($limit > 0) {{
            $query->range(0, $limit);
        }}

        $term_ids = $query->execute();

        if (empty($term_ids)) {{
            echo json_encode([]);
            exit;
        }}

        $terms = $term_storage->loadMultiple($term_ids);
        $results = [];
        $total_terms = count($terms);

        fwrite(STDERR, "Analyzing " . $total_terms . " terms with full node details...\\n");

        // Get field map for taxonomy references
        $field_map = \\Drupal::service('entity_field.manager')->getFieldMapByFieldType('entity_reference');

        // Build list of field tables to check
        $field_tables = [];
        foreach ($field_map['node'] ?? [] as $field_name => $field_info) {{
            $table_name = 'node__' . $field_name;
            $column_name = $field_name . '_target_id';

            try {{
                $database->query("SELECT 1 FROM {{" . $table_name . "}} LIMIT 1")->fetchField();
                $field_tables[] = [
                    'table' => $table_name,
                    'column' => $column_name,
                    'field_name' => $field_name
                ];
            }} catch (\\Exception $e) {{
                // Table doesn't exist - skip
            }}
        }}

        $processed = 0;
        foreach ($terms as $term) {{
            $processed++;
            if ($processed % 50 === 0 || $processed === 1 || $processed === $total_terms) {{
                fwrite(STDERR, "Progress: " . $processed . "/" . $total_terms . " terms\\n");
            }}

            $tid = (int)$term->id();

            // Get term metadata
            $parents = $term_storage->loadParents($tid);
            $parent_name = $parents ? reset($parents)->getName() : null;

            $children = $term_storage->loadChildren($tid);
            $children_names = array_map(function($child) {{ return $child->getName(); }}, $children);

            // Find all nodes using this term
            $nids_found = [];

            // Check taxonomy_index first (fast)
            $taxonomy_index_nids = $database->query(
                "SELECT DISTINCT nid FROM {{taxonomy_index}} WHERE tid = :tid",
                [':tid' => $tid]
            )->fetchCol();

            if ($taxonomy_index_nids) {{
                $nids_found = array_merge($nids_found, $taxonomy_index_nids);
            }}

            // Check field data tables
            foreach ($field_tables as $field_info) {{
                try {{
                    $field_nids = $database->query(
                        "SELECT DISTINCT entity_id FROM {{" . $field_info['table'] . "}} WHERE " . $field_info['column'] . " = :tid",
                        [':tid' => $tid]
                    )->fetchCol();

                    if ($field_nids) {{
                        $nids_found = array_merge($nids_found, $field_nids);
                    }}
                }} catch (\\Exception $e) {{
                    // Query failed - skip
                }}
            }}

            // Get unique NIDs
            $nids_found = array_unique($nids_found);
            $total_content_count = count($nids_found);

            // Load full node details for ALL nodes
            $nodes_data = [];
            if (!empty($nids_found)) {{
                $nodes = $node_storage->loadMultiple($nids_found);
                foreach ($nodes as $node) {{
                    $author = $node->getOwner();

                    // URL and alias
                    $url_alias = '';
                    $canonical_url = '';
                    try {{
                        $url = $node->toUrl('canonical', ['absolute' => FALSE])->toString();
                        $url_alias = $url;
                        $canonical_url = $node->toUrl('canonical', ['absolute' => TRUE])->toString();
                    }} catch (\\Exception $e) {{
                        // Node might not have a URL
                    }}

                    // Taxonomy terms (all vocabularies)
                    $taxonomy_terms = [];
                    foreach ($node->getFields() as $field_name => $field) {{
                        if ($field->getFieldDefinition()->getType() === 'entity_reference' &&
                            $field->getFieldDefinition()->getSetting('target_type') === 'taxonomy_term') {{
                            foreach ($field->referencedEntities() as $t) {{
                                $term_vocab = $t->bundle();
                                $taxonomy_terms[] = $term_vocab . ': ' . $t->getName();
                            }}
                        }}
                    }}

                    // Entity references
                    $entity_refs = [];
                    foreach ($node->getFields() as $field_name => $field) {{
                        if ($field->getFieldDefinition()->getType() === 'entity_reference') {{
                            $target_type = $field->getFieldDefinition()->getSetting('target_type');
                            if ($target_type !== 'taxonomy_term') {{
                                foreach ($field->referencedEntities() as $ref_entity) {{
                                    $entity_refs[] = $target_type . ':' . $ref_entity->id();
                                }}
                            }}
                        }}
                    }}

                    // Metatags
                    $metatag_title = '';
                    $metatag_description = '';
                    if (\\Drupal::moduleHandler()->moduleExists('metatag') && $node->hasField('field_metatag')) {{
                        $metatags = $node->get('field_metatag')->getValue();
                        if (!empty($metatags[0])) {{
                            $metatag_title = $metatags[0]['title'] ?? '';
                            $metatag_description = $metatags[0]['description'] ?? '';
                        }}
                    }}

                    // Revision count
                    $revision_ids = $database->select('node_revision', 'nr')
                        ->fields('nr', ['vid'])
                        ->condition('nid', $node->id())
                        ->execute()
                        ->fetchCol();
                    $revision_count = count($revision_ids);

                    $nodes_data[] = [
                        'nid' => (int) $node->id(),
                        'uuid' => $node->uuid(),
                        'title' => $node->getTitle(),
                        'type' => $node->bundle(),
                        'status' => $node->isPublished() ? 'published' : 'unpublished',
                        'langcode' => $node->language()->getId(),
                        'created' => (int) $node->getCreatedTime(),
                        'changed' => (int) $node->getChangedTime(),
                        'author' => $author->getDisplayName(),
                        'author_uid' => (int) $author->id(),
                        'url_alias' => $url_alias,
                        'canonical_url' => $canonical_url,
                        'promote' => $node->isPromoted(),
                        'sticky' => $node->isSticky(),
                        'taxonomy_terms' => $taxonomy_terms,
                        'entity_references' => $entity_refs,
                        'metatag_title' => $metatag_title,
                        'metatag_description' => $metatag_description,
                        'revision_count' => $revision_count,
                    ];
                }}
            }}

            // Determine if safe to delete
            $safe_to_delete = ($total_content_count === 0 && empty($children_names));

            $warnings = [];
            if ($total_content_count > 0) {{
                $warnings[] = "Used in " . $total_content_count . " content items";
            }}
            if (!empty($children_names)) {{
                $warnings[] = "Has " . count($children_names) . " child terms";
            }}

            $results[] = [
                'tid' => $tid,
                'name' => $term->getName(),
                'description' => $term->getDescription(),
                'vocabulary' => $vocabulary,
                'vocabulary_label' => $vocab->label(),
                'parent' => $parent_name,
                'children' => array_values($children_names),
                'content_count' => $total_content_count,
                'nodes' => $nodes_data,
                'safe_to_delete' => $safe_to_delete,
                'warnings' => $warnings
            ];
        }}

        fwrite(STDERR, "Completed: " . $total_terms . " terms with full node details\\n");
        echo json_encode($results);
        """

        result = run_drush_command(["ev", php_code.strip()], timeout=300)
        return result if result and isinstance(result, list) else None

    except Exception as e:
        logger.debug(f"Could not get terms with full nodes from drush: {e}")
        return None


def _auto_size_columns(sheet, columns):
    """Auto-size columns based on content (with reasonable limits)."""
    for col_idx, col_name in enumerate(columns, start=1):
        # Start with header length
        max_length = len(str(col_name))

        # Sample first 100 rows to determine width
        for row_idx in range(2, min(102, sheet.max_row + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        # Set column width with reasonable limits
        adjusted_width = min(max_length + 2, 60)  # Max 60 chars
        adjusted_width = max(adjusted_width, 10)   # Min 10 chars

        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = adjusted_width
